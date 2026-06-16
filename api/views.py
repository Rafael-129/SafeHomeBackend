import os
from datetime import date, datetime, timedelta
from django.utils import timezone

from django.db.models import Count
from django.db.models.deletion import ProtectedError
from django.db import IntegrityError, transaction
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters, status, viewsets
from rest_framework.decorators import action, api_view
from rest_framework.response import Response
from django.http import HttpResponse, JsonResponse
import io
import csv
from datetime import datetime as _dt

from .models import (
    Configuracion,
    Departamento,
    EventosSistema,
    HistorialAccesos,
    IngresoEventual,
    Notificaciones,
    PerfilAplicacion,
    Scanner,
    SesionesAdmin,
    Usuario,
    UsuarioAdmin,
    Visitante,
    Incidentes,
)
from .retention import RETENCION_DIAS_DEFAULT, purgar_fotos_vencidas
from .serializers import (
    ConfiguracionSerializer,
    DepartamentoSerializer,
    EventosSistemaSerializer,
    HistorialAccesosSerializer,
    IngresoEventualSerializer,
    NotificacionesSerializer,
    PerfilAplicacionSerializer,
    ScannerSerializer,
    SesionesAdminSerializer,
    UsuarioSerializer,
    UsuarioAdminSerializer,
    VisitanteSerializer,
    IncidentesSerializer,
)


@api_view(['GET'])
def api_root(request):
    return Response({
        'message': 'Bienvenido a la API de SafeHome',
        'version': '1.0',
        'endpoints': {
            'admin': '/admin/',
            'api': '/api/',
            'departamentos': '/api/departamentos/',
            'usuarios': '/api/usuarios/',
            'visitantes': '/api/visitantes/',
            'visitantes_frecuentes': '/api/visitantes/frecuentes/',
            'scanner': '/api/scanner/',
            'ingresos_eventuales': '/api/ingresos-eventuales/',
            'historial': '/api/historial/',
            'historial_estadisticas': '/api/historial/estadisticas/?desde=YYYY-MM-DD&hasta=YYYY-MM-DD',
            'perfil_actual': '/api/perfil/actual/',
            'admins': '/api/admins/',
            'sesiones_admin': '/api/sesiones-admin/',
            'eventos_sistema': '/api/eventos-sistema/',
            'configuracion': '/api/configuracion/',
            'notificaciones': '/api/notificaciones/',
            'health': '/api/health/',
        },
    })


@api_view(['GET'])
def health_check(request):
    return Response({
        'status': 'ok',
        'message': 'API funcionando correctamente',
    }, status=status.HTTP_200_OK)


@api_view(['GET', 'POST'])
def generar_reporte(request):
    """Genera y devuelve reportes en CSV/PDF/XLSX según parámetros.
    Parámetros (GET o POST JSON):
    - tipo: Nombre del reporte (opcional, afecta filtros básicos)
    - desde: YYYY-MM-DD
    - hasta: YYYY-MM-DD
    - formato: csv | pdf | xlsx (default csv)
    """
    tipo = request.query_params.get('tipo') or (request.data.get('tipo') if hasattr(request, 'data') else None)
    formato = (request.query_params.get('formato') or (request.data.get('formato') if hasattr(request, 'data') else None) or 'csv').lower()
    desde = request.query_params.get('desde') or (request.data.get('desde') if hasattr(request, 'data') else None)
    hasta = request.query_params.get('hasta') or (request.data.get('hasta') if hasattr(request, 'data') else None)

    qs = HistorialAccesos.objects.select_related('idusuario', 'idvisitante').all()
    try:
        if desde:
            qs = qs.filter(fecha_entrada__gte=_dt.strptime(desde, '%Y-%m-%d').date())
        if hasta:
            qs = qs.filter(fecha_entrada__lte=_dt.strptime(hasta, '%Y-%m-%d').date())
    except ValueError:
        return JsonResponse({'error': 'Formato de fecha inválido. Use YYYY-MM-DD en desde/hasta.'}, status=400)

    # apply basic tipo filters
    if tipo:
        if 'seguridad' in tipo.lower():
            qs = qs.filter(estado__icontains='Deneg')
        if 'visitante' in tipo.lower():
            qs = qs.filter(idvisitante__isnull=False)
        if 'residente' in tipo.lower():
            qs = qs.filter(idusuario__isnull=False)

    serializer = HistorialAccesosSerializer(qs, many=True)
    rows = serializer.data

    filename_base = (tipo or 'reporte').replace(' ', '_')

    # CSV (fast path)
    if formato == 'csv':
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Persona', 'Tipo', 'Accion', 'Hora', 'Fecha', 'Ubicacion', 'Estado'])
        for r in rows:
            persona = None
            if r.get('usuario_info'):
                ui = r['usuario_info']
                persona = f"{ui.get('nombre','')} {ui.get('apellido','')}".strip()
            elif r.get('visitante_info'):
                vi = r['visitante_info']
                persona = f"{vi.get('nombre','')} {vi.get('apellido','')}".strip()
            else:
                persona = 'Desconocido'
            tipo_persona = 'Residente' if r.get('idusuario') else ('Visitante' if r.get('idvisitante') else 'No Identificado')
            estado = 'Exitoso' if str(r.get('estado')) == 'Permitido' else 'Denegado'
            hora = r.get('hora_entrada') or r.get('hora') or '-'
            fecha = r.get('fecha_entrada') or r.get('fecha') or ''
            writer.writerow([persona, tipo_persona, 'Acceso Autorizado' if estado == 'Exitoso' else 'Acceso Denegado', hora, fecha, 'Entrada Principal', estado])
        content = output.getvalue().encode('utf-8')
        resp = HttpResponse(content, content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = f'attachment; filename="{filename_base}.csv"'
        return resp

    # PDF generation (requires reportlab)
    if formato == 'pdf':
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas
        except Exception as e:
            return JsonResponse({'error': 'Falta la dependencia reportlab. Instale: pip install reportlab'}, status=501)

        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        y = 750
        p.setFont('Helvetica-Bold', 12)
        p.drawString(40, y, f'Reporte: {tipo or "Generado"} - {_dt.now().isoformat()}')
        y -= 30
        p.setFont('Helvetica', 10)
        p.drawString(40, y, 'Persona | Tipo | Accion | Hora | Fecha | Estado')
        y -= 18
        for r in rows[:200]:
            persona = r.get('usuario_info') and f"{r['usuario_info'].get('nombre','')} {r['usuario_info'].get('apellido','')}" or (r.get('visitante_info') and f"{r['visitante_info'].get('nombre','')} {r['visitante_info'].get('apellido','')}" or 'Desconocido')
            tipo_persona = 'Residente' if r.get('idusuario') else ('Visitante' if r.get('idvisitante') else 'No Identificado')
            estado = 'Exitoso' if str(r.get('estado')) == 'Permitido' else 'Denegado'
            hora = r.get('hora_entrada') or r.get('hora') or '-'
            fecha = r.get('fecha_entrada') or r.get('fecha') or ''
            text = f"{persona} | {tipo_persona} | {'Acceso Autorizado' if estado=='Exitoso' else 'Acceso Denegado'} | {hora} | {fecha} | {estado}"
            p.drawString(40, y, (text[:120]))
            y -= 14
            if y < 60:
                p.showPage()
                y = 750
        p.save()
        buffer.seek(0)
        resp = HttpResponse(buffer.read(), content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
        return resp

    # XLSX generation (requires openpyxl)
    if formato in ('xlsx', 'xls'):
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except Exception:
            return JsonResponse({'error': 'Falta la dependencia openpyxl. Instale: pip install openpyxl'}, status=501)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Reporte'
        header = ['Persona', 'Tipo', 'Accion', 'Hora', 'Fecha', 'Ubicacion', 'Estado']
        ws.append(header)
        for r in rows:
            persona = None
            if r.get('usuario_info'):
                ui = r['usuario_info']
                persona = f"{ui.get('nombre','')} {ui.get('apellido','')}".strip()
            elif r.get('visitante_info'):
                vi = r['visitante_info']
                persona = f"{vi.get('nombre','')} {vi.get('apellido','')}".strip()
            else:
                persona = 'Desconocido'
            tipo_persona = 'Residente' if r.get('idusuario') else ('Visitante' if r.get('idvisitante') else 'No Identificado')
            estado = 'Exitoso' if str(r.get('estado')) == 'Permitido' else 'Denegado'
            hora = r.get('hora_entrada') or r.get('hora') or '-'
            fecha = r.get('fecha_entrada') or r.get('fecha') or ''
            ws.append([persona, tipo_persona, 'Acceso Autorizado' if estado == 'Exitoso' else 'Acceso Denegado', hora, fecha, 'Entrada Principal', estado])

        for i, _ in enumerate(header, start=1):
            ws.column_dimensions[get_column_letter(i)].width = 20

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        resp = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
        return resp

    return JsonResponse({'error': 'Formato no soportado. Use csv|pdf|xlsx.'}, status=400)


class DepartamentoViewSet(viewsets.ModelViewSet):
    queryset = Departamento.objects.all()
    serializer_class = DepartamentoSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['torre', 'piso', 'numero']
    search_fields = ['codigo', 'torre']
    ordering_fields = ['torre', 'piso', 'numero']
    ordering = ['torre', 'piso', 'numero']


class UsuarioViewSet(viewsets.ModelViewSet):
    queryset = Usuario.objects.all()
    serializer_class = UsuarioSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['iddepartamento', 'dni']
    search_fields = ['nombre', 'apellido', 'dni', 'correo']
    ordering_fields = ['nombre', 'apellido', 'dni']
    ordering = ['nombre', 'apellido']

    @action(detail=False, methods=['get'])
    def buscar_por_dni(self, request):
        dni = request.query_params.get('dni', None)
        if dni:
            try:
                usuario = Usuario.objects.get(dni=dni)
                serializer = self.get_serializer(usuario)
                return Response(serializer.data)
            except Usuario.DoesNotExist:
                return Response({'error': 'Usuario no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        return Response({'error': 'DNI no proporcionado'}, status=status.HTTP_400_BAD_REQUEST)


class VisitanteViewSet(viewsets.ModelViewSet):
    queryset = Visitante.objects.all()
    serializer_class = VisitanteSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['dni', 'iddepartamento', 'fecha_visita']
    search_fields = ['nombre', 'apellido', 'dni', 'motivo']
    ordering_fields = ['nombre', 'apellido', 'fecha_visita']
    ordering = ['-fecha_visita']

    @action(detail=False, methods=['get'])


    def hoy(self, request):
        hoy = date.today()
        visitantes = self.queryset.filter(fecha_visita=hoy)
        serializer = self.get_serializer(visitantes, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def frecuentes(self, request):
        top = int(request.query_params.get('top', 10))
        rows = (
            Visitante.objects.values('dni', 'nombre', 'apellido')
            .annotate(total_visitas=Count('idvisitante'))
            .order_by('-total_visitas', 'apellido', 'nombre')[:top]
        )
        return Response(list(rows), status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def purgar_fotos(self, request):
        """Borra las fotos de visitantes vencidas (politica de retencion).

        Protegido opcionalmente: si existe la env PURGE_TOKEN, exige el header
        X-Purge-Token con el mismo valor.
        """
        token_req = os.getenv('PURGE_TOKEN')
        if token_req and request.headers.get('X-Purge-Token') != token_req:
            return Response({'error': 'No autorizado'}, status=status.HTTP_401_UNAUTHORIZED)

        dias = request.data.get('dias') or os.getenv('RETENCION_FOTOS_DIAS', RETENCION_DIAS_DEFAULT)
        try:
            resultado = purgar_fotos_vencidas(dias)
        except (TypeError, ValueError):
            return Response({'error': 'Parametro dias invalido.'}, status=status.HTTP_400_BAD_REQUEST)
        return Response(resultado, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    def finalizar(self, request, pk=None):
        visitante = self.get_object()
        # preserve original visit timestamps if present
        original_fecha = visitante.fecha_visita
        original_hora = visitante.hora_visita

        finalizado_en = timezone.now() - timedelta(hours=3)
        visitante.fecha_visita = finalizado_en.date()
        visitante.hora_visita = finalizado_en.time().replace(microsecond=0)
        visitante.save(update_fields=['fecha_visita', 'hora_visita'])

        # Update or create historial entry: set hora_salida for active access records
        try:
            historial = HistorialAccesos.objects.filter(idvisitante=visitante, hora_salida__isnull=True).order_by('-fecha_entrada', '-hora_entrada').first()
            if historial:
                historial.hora_salida = finalizado_en.time().replace(microsecond=0)
                historial.save(update_fields=['hora_salida'])
            else:
                HistorialAccesos.objects.create(
                    idvisitante=visitante,
                    fecha_entrada=original_fecha if original_fecha else finalizado_en.date(),
                    hora_entrada=original_hora if original_hora else finalizado_en.time().replace(microsecond=0),
                    hora_salida=finalizado_en.time().replace(microsecond=0),
                    estado='Permitido',
                )
        except Exception:
            # don't block finalize on historial issues; return visitante update regardless
            pass

        serializer = self.get_serializer(visitante)
        return Response(
            {
                'message': 'Visita finalizada correctamente. Hora de salida registrada en historial cuando fue posible.',
                'visitante': serializer.data,
            },
            status=status.HTTP_200_OK,
        )

    def destroy(self, request, *args, **kwargs):
        try:
            return super().destroy(request, *args, **kwargs)
        except (ProtectedError, IntegrityError):
            return Response(
                {
                    'error': 'No se puede eliminar este visitante porque tiene registros relacionados. Use /visitantes/{id}/finalizar/.',
                },
                status=status.HTTP_409_CONFLICT,
            )


class ScannerViewSet(viewsets.ModelViewSet):
    queryset = Scanner.objects.all()
    serializer_class = ScannerSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['tipo_persona', 'idusuario', 'idvisitante']
    search_fields = ['tipo_persona']
    ordering_fields = ['fecha']
    ordering = ['-fecha']

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        validated = serializer.validated_data
        tipo_persona = request.data.get('tipo_persona') or validated.get('tipo_persona')
        dni = request.data.get('dni')

        usuario = validated.get('idusuario')
        visitante = validated.get('idvisitante')

        if not usuario and not visitante and dni:
            if tipo_persona == 'residente':
                usuario = Usuario.objects.filter(dni=dni).first()
            elif tipo_persona == 'visitante':
                visitante = Visitante.objects.filter(dni=dni).order_by('-fecha_visita', '-hora_visita').first()
            else:
                usuario = Usuario.objects.filter(dni=dni).first()
                if not usuario:
                    visitante = Visitante.objects.filter(dni=dni).order_by('-fecha_visita', '-hora_visita').first()

        if not usuario and not visitante:
            # Desconocido: persistimos el escaneo (con su foto) para dejar evidencia
            # y poder enlazarlo a un acceso denegado en el historial.
            scanner = serializer.save(
                idusuario=None,
                idvisitante=None,
                tipo_persona='desconocido',
            )
            output = self.get_serializer(scanner).data
            output['autorizado'] = False
            output['mensaje'] = 'No se encontro coincidencia. Acceso denegado.'
            return Response(output, status=status.HTTP_201_CREATED)

        scanner = serializer.save(
            idusuario=usuario,
            idvisitante=visitante,
            tipo_persona='residente' if usuario else 'visitante',
            foto_capturada=None,  # reconocidos: se muestra su foto enrolada, no se guarda captura
        )

        output = self.get_serializer(scanner).data
        output['autorizado'] = True
        output['mensaje'] = 'Persona identificada correctamente.'
        return Response(output, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'])
    def recientes(self, request):
        escaneos = self.queryset.order_by('-fecha')[:50]
        serializer = self.get_serializer(escaneos, many=True)
        return Response(serializer.data)


class IngresoEventualViewSet(viewsets.ModelViewSet):
    queryset = IngresoEventual.objects.select_related('iddepartamento').all()
    serializer_class = IngresoEventualSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['dni', 'iddepartamento']
    search_fields = ['nombre', 'apellido', 'dni', 'motivo']
    ordering_fields = ['fecha', 'nombre', 'apellido']
    ordering = ['-fecha']

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        with transaction.atomic():
            eventual = serializer.save()
            ahora = timezone.localtime()
            historial = HistorialAccesos.objects.create(
                ideventual=eventual,
                fecha_entrada=ahora.date(),
                hora_entrada=ahora.time().replace(microsecond=0),
                estado='Permitido',
            )

        output = self.get_serializer(eventual).data
        output['autorizado'] = True
        output['mensaje'] = 'Ingreso eventual registrado correctamente.'
        output['idhistorial'] = historial.idhistorial
        return Response(output, status=status.HTTP_201_CREATED)


class HistorialAccesosViewSet(viewsets.ModelViewSet):
    queryset = HistorialAccesos.objects.select_related(
        'idusuario',
        'idvisitante',
        'idscanner',
        'ideventual',
        'idusuario__iddepartamento',
        'idvisitante__iddepartamento',
        'ideventual__iddepartamento',
    ).all()
    serializer_class = HistorialAccesosSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['estado', 'fecha_entrada', 'idusuario', 'idvisitante']
    search_fields = ['estado']
    ordering_fields = ['fecha_entrada', 'hora_entrada']
    ordering = ['-fecha_entrada', '-hora_entrada']

    @action(detail=False, methods=['get'])
    def hoy(self, request):
        hoy = date.today()
        accesos = self.queryset.filter(fecha_entrada=hoy)
        serializer = self.get_serializer(accesos, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def activos(self, request):
        accesos = self.queryset.filter(hora_salida__isnull=True, estado='Permitido')
        serializer = self.get_serializer(accesos, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def estadisticas(self, request):
        desde = request.query_params.get('desde')
        hasta = request.query_params.get('hasta')

        qs = self.queryset
        try:
            if desde:
                qs = qs.filter(fecha_entrada__gte=datetime.strptime(desde, '%Y-%m-%d').date())
            if hasta:
                qs = qs.filter(fecha_entrada__lte=datetime.strptime(hasta, '%Y-%m-%d').date())
        except ValueError:
            return Response(
                {'error': 'Formato de fecha inválido. Use YYYY-MM-DD en desde/hasta.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        total = qs.count()
        autorizados = qs.filter(estado='Permitido').count()
        denegados = qs.filter(estado='Denegado').count()
        residentes = qs.filter(idusuario__isnull=False).count()
        visitantes = qs.filter(idvisitante__isnull=False).count()
        por_estado = list(qs.values('estado').annotate(total=Count('idhistorial')).order_by('estado'))

        return Response(
            {
                'totalAccesos': total,
                'autorizados': autorizados,
                'denegados': denegados,
                'residentes': residentes,
                'visitantes': visitantes,
                'porEstado': por_estado,
                'desde': desde,
                'hasta': hasta,
            },
            status=status.HTTP_200_OK,
        )


class PerfilAplicacionViewSet(viewsets.ModelViewSet):
    queryset = PerfilAplicacion.objects.all()
    serializer_class = PerfilAplicacionSerializer

    @action(detail=False, methods=['get', 'put', 'patch'])
    def actual(self, request):
        perfil, _ = PerfilAplicacion.objects.get_or_create(
            idperfil=1,
            defaults={
                'nombre_aplicacion': 'SafeHome Scanner',
                'descripcion': 'Sistema de control de acceso para condominio',
                'version': '1.0.0',
                'permitir_registro_sin_foto': True,
                'politica_foto_requerida': False,
            },
        )

        if request.method in ['PUT', 'PATCH']:
            partial = request.method == 'PATCH'
            serializer = self.get_serializer(perfil, data=request.data, partial=partial)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            return Response(serializer.data)

        serializer = self.get_serializer(perfil)
        return Response(serializer.data)


class UsuarioAdminViewSet(viewsets.ModelViewSet):
    queryset = UsuarioAdmin.objects.all()
    serializer_class = UsuarioAdminSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['rol', 'activo', 'autenticacion_2fa']
    search_fields = ['username', 'nombre_completo', 'email']
    ordering_fields = ['username', 'created_at', 'ultimo_acceso']
    ordering = ['username']


class SesionesAdminViewSet(viewsets.ModelViewSet):
    queryset = SesionesAdmin.objects.select_related('idadmin').all()
    serializer_class = SesionesAdminSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['activa', 'idadmin']
    search_fields = ['token', 'idadmin__username', 'ip_address']
    ordering_fields = ['fecha_inicio', 'fecha_expiracion']
    ordering = ['-fecha_inicio']


class EventosSistemaViewSet(viewsets.ModelViewSet):
    queryset = EventosSistema.objects.select_related('idadmin').all()
    serializer_class = EventosSistemaSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['tipo', 'nivel', 'idadmin']
    search_fields = ['tipo', 'descripcion', 'ip_address', 'idadmin__username']
    ordering_fields = ['fecha', 'nivel']
    ordering = ['-fecha']


class IncidentesViewSet(viewsets.ModelViewSet):
    queryset = Incidentes.objects.select_related('idscanner').all()
    serializer_class = IncidentesSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['gravedad', 'resuelto', 'idscanner']
    search_fields = ['tipo', 'descripcion', 'observaciones']
    ordering_fields = ['fecha', 'gravedad']
    ordering = ['-fecha']

    @action(detail=False, methods=['get'])
    def recientes(self, request):
        qs = self.queryset.order_by('-fecha')[:50]
        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)


class ConfiguracionViewSet(viewsets.ModelViewSet):
    queryset = Configuracion.objects.all()
    serializer_class = ConfiguracionSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['tipo', 'parametro']
    search_fields = ['parametro', 'descripcion', 'valor']
    ordering_fields = ['tipo', 'parametro', 'ultima_modificacion']
    ordering = ['tipo', 'parametro']


class NotificacionesViewSet(viewsets.ModelViewSet):
    queryset = Notificaciones.objects.select_related('idusuario').all()
    serializer_class = NotificacionesSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['tipo', 'leida', 'idusuario']
    search_fields = ['titulo', 'mensaje', 'tipo', 'idusuario__nombre', 'idusuario__apellido']
    ordering_fields = ['fecha', 'tipo']
    ordering = ['-fecha']

    @action(detail=False, methods=['get'])
    def no_leidas(self, request):
        qs = self.queryset.filter(leida=False)
        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    def marcar_leida(self, request, pk=None):
        notificacion = self.get_object()
        notificacion.leida = True
        notificacion.save(update_fields=['leida'])
        serializer = self.get_serializer(notificacion)
        return Response(serializer.data, status=status.HTTP_200_OK)
